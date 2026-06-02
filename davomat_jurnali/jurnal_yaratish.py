# -*- coding: utf-8 -*-
"""
O'quvchilar ro'yxati va darslarga davomat jurnalini (Excel) yaratadi.
Belgilar:  +  (qatnashdi),  -  (qatnashmadi),  0  (sababli/kech qoldi)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------ sozlamalar
OQUVCHILAR_SONI = 20     # o'quvchilar uchun qatorlar soni
DARSLAR_SONI    = 30     # darslar soni (Dars 1 ... Dars 30)
FAYL_NOMI       = "Davomat_jurnali.xlsx"

# ------------------------------------------------------------------ ranglar / stillar
QORA      = "FF000000"
KOK       = "1F4E78"   # sarlavha foni (to'q ko'k)
OCHIQ_KOK = "DDEBF7"   # ustun sarlavhalari foni
OQ        = "FFFFFFFF"
KULRANG   = "F2F2F2"   # juft qatorlar foni
YASHIL    = "C6EFCE"   # +  fon
QIZIL     = "FFC7CE"   # -  fon
SARIQ     = "FFEB9C"   # 0  fon

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

bold_white = Font(name="Calibri", size=11, bold=True, color=OQ)
bold_dark  = Font(name="Calibri", size=11, bold=True, color="1F4E78")
normal     = Font(name="Calibri", size=11, color=QORA)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center")

title_fill  = PatternFill("solid", fgColor=KOK)
header_fill = PatternFill("solid", fgColor=OCHIQ_KOK)
gray_fill   = PatternFill("solid", fgColor=KULRANG)

wb = Workbook()

# ====================================================================
# 1-LIST:  O'QUVCHILAR RO'YXATI
# ====================================================================
ws1 = wb.active
ws1.title = "O'quvchilar"

# Sarlavha
ws1.merge_cells("A1:D1")
c = ws1["A1"]
c.value = "O'QUVCHILAR RO'YXATI"
c.font = Font(name="Calibri", size=14, bold=True, color=OQ)
c.fill = title_fill
c.alignment = center
ws1.row_dimensions[1].height = 26

# Ustun sarlavhalari
ustunlar1 = ["№", "F.I.SH.", "Telefon raqami", "Izoh"]
for j, nom in enumerate(ustunlar1, start=1):
    cell = ws1.cell(row=2, column=j, value=nom)
    cell.font = bold_dark
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
ws1.row_dimensions[2].height = 20

# O'quvchilar qatorlari (raqamlangan, ismlar bo'sh)
for i in range(OQUVCHILAR_SONI):
    r = 3 + i
    ws1.cell(row=r, column=1, value=i + 1).alignment = center
    for j in range(1, 5):
        cell = ws1.cell(row=r, column=j)
        cell.border = border
        cell.font = normal
        if r % 2 == 0:
            cell.fill = gray_fill
    ws1.row_dimensions[r].height = 18

# Ustun kengliklari
ws1.column_dimensions["A"].width = 6
ws1.column_dimensions["B"].width = 32
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 22
ws1.freeze_panes = "A3"

# ====================================================================
# 2-LIST:  DAVOMAT
# ====================================================================
ws2 = wb.create_sheet("Davomat")

ILK_DARS_USTUN = 3                      # C ustunidan boshlab darslar
oxirgi_dars_ust = ILK_DARS_USTUN + DARSLAR_SONI - 1
jami_p_ust   = oxirgi_dars_ust + 1      # Jami +
jami_m_ust   = oxirgi_dars_ust + 2      # Jami -
jami_n_ust   = oxirgi_dars_ust + 3      # Jami 0
foiz_ust     = oxirgi_dars_ust + 4      # Davomat %
oxirgi_ustun = foiz_ust

# --- Umumiy sarlavha (1-qator)
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=oxirgi_ustun)
c = ws2.cell(row=1, column=1, value="DARSLARGA DAVOMAT JURNALI")
c.font = Font(name="Calibri", size=14, bold=True, color=OQ)
c.fill = title_fill
c.alignment = center
ws2.row_dimensions[1].height = 26

# --- № va F.I.SH. (2-3 qatorlar birlashtirilgan)
ws2.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
ws2.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
hn = ws2.cell(row=2, column=1, value="№")
hf = ws2.cell(row=2, column=2, value="F.I.SH.")
for cell in (hn, hf):
    cell.font = bold_dark
    cell.fill = header_fill
    cell.alignment = center

# --- Darslar sarlavhasi (2-qator) + Sana qatori (3-qator, bo'sh)
for k in range(DARSLAR_SONI):
    col = ILK_DARS_USTUN + k
    # 2-qator: "1-dars", "2-dars" ...
    dc = ws2.cell(row=2, column=col, value=f"{k + 1}-dars")
    dc.font = bold_dark
    dc.fill = header_fill
    dc.alignment = center
    # 3-qator: sana uchun BO'SH joy (foydalanuvchi o'zi kiritadi)
    sc = ws2.cell(row=3, column=col)
    sc.fill = PatternFill("solid", fgColor=OQ)
    sc.alignment = center
    sc.font = Font(name="Calibri", size=9, italic=True, color="808080")
    ws2.column_dimensions[get_column_letter(col)].width = 6

# 3-qatorning chap tomonida ko'rsatma
sana_label = ws2.cell(row=3, column=2)
# B ustuni 2-3 birlashtirilgani uchun label qo'ya olmaymiz -> A ustunida bo'sh.
# Buning o'rniga izohni pastga yozamiz.

# --- Yig'indi (Jami) ustunlari sarlavhasi
jami_sarlavha = {
    jami_p_ust: "Jami +",
    jami_m_ust: "Jami -",
    jami_n_ust: "Jami 0",
    foiz_ust:   "Davomat %",
}
for col, nom in jami_sarlavha.items():
    ws2.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
    cc = ws2.cell(row=2, column=col, value=nom)
    cc.font = bold_dark
    cc.fill = PatternFill("solid", fgColor="E2EFDA")
    cc.alignment = center
    ws2.column_dimensions[get_column_letter(col)].width = 11

ws2.row_dimensions[2].height = 20
ws2.row_dimensions[3].height = 18

# --- O'quvchilar qatorlari
ILK_QATOR = 4
for i in range(OQUVCHILAR_SONI):
    r = ILK_QATOR + i
    ws2.cell(row=r, column=1, value=i + 1).alignment = center
    # Ism 1-listdan avtomatik tortib olinadi
    nom_kataki = ws2.cell(row=r, column=2)
    nom_kataki.value = f"=IF('O''quvchilar'!B{3 + i}=\"\",\"\",'O''quvchilar'!B{3 + i})"
    nom_kataki.alignment = left
    nom_kataki.font = normal

    # darslar (bo'sh kataklar)
    for k in range(DARSLAR_SONI):
        col = ILK_DARS_USTUN + k
        cell = ws2.cell(row=r, column=col)
        cell.alignment = center
        cell.font = normal

    # Jami formulalar
    dars_boshi = get_column_letter(ILK_DARS_USTUN) + str(r)
    dars_oxiri = get_column_letter(oxirgi_dars_ust) + str(r)
    rng = f"{dars_boshi}:{dars_oxiri}"
    ws2.cell(row=r, column=jami_p_ust, value=f'=COUNTIF({rng},"+")').alignment = center
    ws2.cell(row=r, column=jami_m_ust, value=f'=COUNTIF({rng},"-")').alignment = center
    ws2.cell(row=r, column=jami_n_ust, value=f'=COUNTIF({rng},"0")').alignment = center
    # Davomat foizi: + larning belgilangan kataklarga nisbati
    p = get_column_letter(jami_p_ust) + str(r)
    m = get_column_letter(jami_m_ust) + str(r)
    n = get_column_letter(jami_n_ust) + str(r)
    foiz = ws2.cell(
        row=r, column=foiz_ust,
        value=f'=IF(({p}+{m}+{n})=0,"",ROUND({p}/({p}+{m}+{n})*100,0))'
    )
    foiz.alignment = center
    foiz.number_format = '0"%"'

    # barcha kataklarga ramka + juft qator foni
    for col in range(1, oxirgi_ustun + 1):
        cell = ws2.cell(row=r, column=col)
        cell.border = border
        if r % 2 == 1 and col <= oxirgi_dars_ust:
            cell.fill = gray_fill
    ws2.row_dimensions[r].height = 18

# 2-3 sarlavha kataklariga ramka
for r in (2, 3):
    for col in range(1, oxirgi_ustun + 1):
        ws2.cell(row=r, column=col).border = border

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 30
ws2.freeze_panes = "C4"   # № va F.I.SH. hamda sarlavhalar qotirilgan

# --- Faqat +, -, 0 kiritishga ruxsat (ma'lumot tekshiruvi)
oxirgi_qator = ILK_QATOR + OQUVCHILAR_SONI - 1
dv = DataValidation(
    type="list",
    formula1='"+,-,0"',
    allow_blank=True,
    showErrorMessage=True,
)
dv.error = "Faqat  + , -  yoki  0  kiriting!"
dv.errorTitle = "Noto'g'ri belgi"
dv.prompt = "+ qatnashdi,  - qatnashmadi,  0 sababli"
dv.promptTitle = "Davomat belgisi"
ws2.add_data_validation(dv)
dv.add(
    f"{get_column_letter(ILK_DARS_USTUN)}{ILK_QATOR}:"
    f"{get_column_letter(oxirgi_dars_ust)}{oxirgi_qator}"
)

# --- Shartli formatlash: + yashil, - qizil, 0 sariq
dars_diapazoni = (
    f"{get_column_letter(ILK_DARS_USTUN)}{ILK_QATOR}:"
    f"{get_column_letter(oxirgi_dars_ust)}{oxirgi_qator}"
)
ws2.conditional_formatting.add(
    dars_diapazoni,
    CellIsRule(operator="equal", formula=['"+"'],
               fill=PatternFill("solid", fgColor=YASHIL),
               font=Font(color="006100", bold=True))
)
ws2.conditional_formatting.add(
    dars_diapazoni,
    CellIsRule(operator="equal", formula=['"-"'],
               fill=PatternFill("solid", fgColor=QIZIL),
               font=Font(color="9C0006", bold=True))
)
ws2.conditional_formatting.add(
    dars_diapazoni,
    CellIsRule(operator="equal", formula=['"0"'],
               fill=PatternFill("solid", fgColor=SARIQ),
               font=Font(color="9C5700", bold=True))
)

# --- Izoh (legend) jadval ostida
izoh_qator = oxirgi_qator + 2
izohlar = [
    ("BELGILAR:", ""),
    ("+", "qatnashdi"),
    ("-", "qatnashmadi"),
    ("0", "sababli / kech qoldi"),
    ("3-qator", "har bir darsning sanasini o'zingiz kiritasiz"),
]
ws2.cell(row=izoh_qator, column=2, value="IZOH").font = bold_dark
for idx, (belgi, mat) in enumerate(izohlar[1:], start=1):
    r = izoh_qator + idx
    b = ws2.cell(row=r, column=2, value=belgi)
    b.font = Font(bold=True)
    b.alignment = Alignment(horizontal="center")
    ws2.cell(row=r, column=ILK_DARS_USTUN, value=mat).font = normal

wb.save(FAYL_NOMI)
print(f"Tayyor: {FAYL_NOMI}")
print(f"  - O'quvchilar: {OQUVCHILAR_SONI} qator")
print(f"  - Darslar: {DARSLAR_SONI} ta")
