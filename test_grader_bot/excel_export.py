"""
Excel export module - generates .xlsx result files using openpyxl.
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def generate_results_excel(results: list) -> BytesIO:
    """
    Generate an Excel file with student results.

    Args:
        results: List of dicts with keys:
            - name: Student first name
            - surname: Student surname
            - score: Number of correct answers
            - total: Total number of questions
            - answers: List of detected answers (for reference)
            - student_number: Student number (optional, int or None)

    Returns:
        BytesIO buffer containing the .xlsx file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Natijalar"

    # Header row
    headers = ["#", "Raqam", "Ism", "Familiya", "Ball", "Foiz", "Javoblar"]
    header_font = Font(bold=True, size=12)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Set column widths
    ws.column_dimensions["A"].width = 6   # #
    ws.column_dimensions["B"].width = 10  # Raqam
    ws.column_dimensions["C"].width = 20  # Ism
    ws.column_dimensions["D"].width = 20  # Familiya
    ws.column_dimensions["E"].width = 10  # Ball
    ws.column_dimensions["F"].width = 10  # Foiz
    ws.column_dimensions["G"].width = 40  # Javoblar

    # Sort results by score descending
    sorted_results = sorted(results, key=lambda x: x.get("score", 0), reverse=True)

    # Data rows
    for rank, student in enumerate(sorted_results, 1):
        student_number = student.get("student_number", None)
        name = student.get("name", "")
        surname = student.get("surname", "")
        score = student.get("score", 0)
        total = student.get("total", 1)
        percentage = round((score / total) * 100, 1) if total > 0 else 0
        answers = student.get("answers", [])
        answers_str = "".join(str(a) for a in answers) if answers else ""

        row = rank + 1
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=student_number if student_number is not None else "-")
        ws.cell(row=row, column=3, value=name)
        ws.cell(row=row, column=4, value=surname)
        ws.cell(row=row, column=5, value=f"{score}/{total}")
        ws.cell(row=row, column=6, value=f"{percentage}%")
        ws.cell(row=row, column=7, value=answers_str)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
